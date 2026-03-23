"""split manifest 读写与数据审计工具。

这个模块是当前主线的数据“事实层”：
- 它把 XLSX 中的行解析成统一样本项；
- 为每条样本补齐 speaker、文本泄漏提示、文件路径、split 等元数据；
- 最后把这些信息固化成 manifest，供训练和推理共享。

这样做的目的，是把“数据事实”和“模型逻辑”分开，减少实验口径漂移。
"""

from __future__ import annotations

import hashlib
import json
import math
import random
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from text_policy_utils import (
    build_prompt_group_id,
    build_prompt_group_text,
    derive_cue_severity,
    mask_emotion_cues,
    normalize_text_for_grouping,
)


CN_TO_EN = {
    "愤怒": "angry",
    "厌恶": "disgusted",
    "恐惧": "fear",
    "高兴": "happy",
    "快乐": "happy",
    "开心": "happy",
    "中性": "neutral",
    "悲伤": "sad",
    "惊讶": "surprise",
}

EMOTIONS = ["angry", "disgusted", "fear", "happy", "neutral", "sad", "surprise"]

SPEAKER_LABELS_EN = {
    "A": {"surprise", "fear", "angry"},
    "B": {"neutral", "disgusted"},
    "C": {"happy", "sad"},
}

EMOTION_WORDS_EN = ["happy", "angry", "sad", "fear", "disgust", "disgusted", "surprise", "neutral"]
EMOTION_WORDS_CN = ["高兴", "开心", "快乐", "愤怒", "生气", "悲伤", "伤心", "恐惧", "害怕", "厌恶", "惊讶", "中性"]
EMOTION_DESC_CN = [
    "很开心",
    "非常开心",
    "特别开心",
    "很愤怒",
    "非常愤怒",
    "特别愤怒",
    "很生气",
    "非常生气",
    "特别生气",
    "很悲伤",
    "非常悲伤",
    "很害怕",
    "非常害怕",
    "很恐惧",
    "非常恐惧",
    "很惊讶",
    "非常惊讶",
    "很厌恶",
    "非常厌恶",
]

TASK_MODES = ("confounded_7way", "within_speaker")


def normalize_seq(raw: Any) -> str:
    """把各种形态的序号字段归一成稳定字符串。

    XLSX 里的序号经常会出现 `1`、`1.0`、`"1"`、空字符串等混杂情况。
    这个函数的目标不是保留原始格式，而是尽量把“语义上同一个样本 ID”
    归并到一个统一表示上，便于后续查路径、对齐缓存和 manifest。
    """

    if raw is None or isinstance(raw, bool):
        return ""
    if isinstance(raw, float) and math.isnan(raw):
        return ""
    if isinstance(raw, (int, float)):
        try:
            v = float(raw)
        except Exception:
            return str(raw).strip()
        if math.isnan(v):
            return ""
        if abs(v - round(v)) < 1e-6:
            return str(int(round(v)))
        return str(raw).strip()

    try:
        v = float(str(raw))
        if not math.isnan(v) and abs(v - round(v)) < 1e-6:
            return str(int(round(v)))
    except Exception:
        pass

    s = str(raw).strip()
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        return str(int(s[:-2]))
    if s.isdigit():
        return str(int(s))
    return s


def parse_intensity(raw: Any) -> float | None:
    """把强度字段解析成浮点数；无效值统一返回 `None`。"""

    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, float) and math.isnan(raw):
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        v = float(s)
    except Exception:
        return None
    if math.isnan(v) or math.isinf(v):
        return None
    return float(v)


def resolve_label_en(raw: Any) -> str | None:
    """把中文标签或英文标签统一映射到主工程使用的英文类别名。"""

    if raw is None:
        return None
    label = CN_TO_EN.get(str(raw).strip(), str(raw).strip().lower()).strip().lower()
    return label if label in EMOTIONS else None


def infer_speaker_id(label_en: str | None) -> str:
    """根据当前已知规则，从情绪标签反推出 speaker ID。

    这是一个带有先验假设的推断，不是数据集中显式给出的字段。
    它主要用于审计当前数据是否存在严重的 speaker confound。
    """

    if not label_en:
        return "UNKNOWN"
    for speaker_id, labels in SPEAKER_LABELS_EN.items():
        if label_en in labels:
            return speaker_id
    return "UNKNOWN"


def _ensure_text_control_fields(item: dict[str, Any]) -> dict[str, Any]:
    """Backfill prompt/text-control fields for older manifests.

    Scientific evaluation may run against manifests created before these fields
    were introduced. Rebuilding a 2k-sample manifest is cheap, but rebuilding on
    a remote machine is still unnecessary churn, so we derive missing fields
    lazily when they are absent.
    """

    enriched = dict(item)
    label_raw = str(enriched.get("label_raw") or "")
    label_en = str(enriched.get("label_en") or "")
    mn = str(enriched.get("mn") or "")
    zh = str(enriched.get("zh") or "")
    cue_details = enriched.get("text_cue_details") or {}

    if not enriched.get("masked_mn"):
        enriched["masked_mn"] = mask_emotion_cues(mn, label_raw=label_raw, label_en=label_en)
    if not enriched.get("masked_zh"):
        enriched["masked_zh"] = mask_emotion_cues(zh, label_raw=label_raw, label_en=label_en)
    if not enriched.get("normalized_mn"):
        enriched["normalized_mn"] = normalize_text_for_grouping(mn)
    if not enriched.get("normalized_zh"):
        enriched["normalized_zh"] = normalize_text_for_grouping(zh)
    if not enriched.get("prompt_group_text"):
        enriched["prompt_group_text"] = build_prompt_group_text(mn, zh, label_raw=label_raw, label_en=label_en)
    if not enriched.get("prompt_group_id"):
        enriched["prompt_group_id"] = build_prompt_group_id(mn, zh, label_raw=label_raw, label_en=label_en)
    if not enriched.get("cue_severity"):
        enriched["cue_severity"] = derive_cue_severity(cue_details)
    return enriched


def resolve_task_mode(task_mode: str | None) -> str:
    """规范化任务模式字符串。"""

    mode = str(task_mode or "confounded_7way").strip().lower()
    if mode not in TASK_MODES:
        raise RuntimeError(f"Unsupported task mode: {task_mode}")
    return mode


def resolve_task_label_names(task_mode: str | None = "confounded_7way", speaker_id: str | None = None) -> list[str]:
    """返回当前任务模式对应的标签集合。

    - `confounded_7way`: 使用全局 7 类标签。
    - `within_speaker`: 只保留某个 speaker 实际覆盖的标签，并按 `EMOTIONS`
      的全局顺序输出，避免不同脚本各自定义局部类别顺序。
    """

    mode = resolve_task_mode(task_mode)
    if mode == "confounded_7way":
        return list(EMOTIONS)

    speaker = str(speaker_id or "").strip().upper()
    if speaker not in SPEAKER_LABELS_EN:
        raise RuntimeError("within_speaker mode requires --speaker-id in {A,B,C}")
    speaker_labels = SPEAKER_LABELS_EN[speaker]
    return [label for label in EMOTIONS if label in speaker_labels]


def map_label_to_task_index(
    label_en: str | None,
    task_mode: str | None = "confounded_7way",
    speaker_id: str | None = None,
) -> int | None:
    """把全局标签映射到当前任务模式下的局部标签索引。"""

    if label_en is None:
        return None
    label_names = resolve_task_label_names(task_mode, speaker_id)
    if label_en not in label_names:
        return None
    return int(label_names.index(label_en))


def filter_manifest_items_for_task(
    items: list[dict[str, Any]],
    task_mode: str | None = "confounded_7way",
    speaker_id: str | None = None,
) -> list[dict[str, Any]]:
    """按任务模式过滤 manifest 条目。"""

    mode = resolve_task_mode(task_mode)
    if mode == "confounded_7way":
        return list(items)

    speaker = str(speaker_id or "").strip().upper()
    label_names = set(resolve_task_label_names(mode, speaker))
    filtered: list[dict[str, Any]] = []
    for item in items:
        if str(item.get("speaker_id", "UNKNOWN")).strip().upper() != speaker:
            continue
        if str(item.get("label_en", "")) not in label_names:
            continue
        filtered.append(item)
    return filtered


def build_validity_summary(
    manifest_summary: dict[str, Any] | None,
    task_mode: str | None = "confounded_7way",
    speaker_id: str | None = None,
) -> dict[str, Any]:
    """把 manifest 审计结果翻译成实验结果可直接消费的有效性摘要。"""

    summary = manifest_summary or {}
    mode = resolve_task_mode(task_mode)
    label_names = resolve_task_label_names(mode, speaker_id)
    fatal_confound = bool(summary.get("fatal_confound", False))
    speaker_group_split_feasible = bool(summary.get("speaker_group_split_feasible", False))

    if mode == "within_speaker":
        speaker = str(speaker_id or "").strip().upper()
        claim_scope = f"within_speaker_emotion_discrimination:{speaker}"
        recommended = (
            f"Current result is scientifically usable only as within-speaker emotion discrimination for speaker {speaker}. "
            "Do not extrapolate it to cross-speaker 7-way generalization."
        )
        scientific_valid = True
    else:
        claim_scope = "multimodal_confounded_7way_benchmark" if fatal_confound else "multimodal_7way_benchmark"
        recommended = (
            "Current result must be interpreted as a confounded 7-way benchmark, not as emotion generalization."
            if fatal_confound
            else "Current result may be interpreted as a 7-way benchmark under the manifest's split assumptions."
        )
        scientific_valid = not fatal_confound

    return {
        "task_mode": mode,
        "speaker_id": str(speaker_id).strip().upper() if speaker_id is not None and str(speaker_id).strip() else None,
        "label_names": label_names,
        "fatal_confound": fatal_confound,
        "speaker_group_split_feasible": speaker_group_split_feasible,
        "scientific_validity": bool(scientific_valid),
        "claim_scope": claim_scope,
        "recommended_interpretation": recommended,
    }


def _contains_keyword(text: str, keywords: list[str], *, lowercase: bool) -> bool:
    """判断文本是否包含任一关键词。"""

    if lowercase:
        haystack = text.lower()
        return any(kw.lower() in haystack for kw in keywords)
    return any(kw in text for kw in keywords)


def detect_text_cue_flags(mn_text: str, zh_text: str, label_raw: str) -> dict[str, bool]:
    """检测文本里是否包含潜在的情绪泄漏线索。

    返回值是一个布尔字典，而不是单一标志位，原因是后续既要判断
    “是否有问题”，也要知道“问题具体来自哪一种文本线索”。
    """

    text = " ".join(x for x in [mn_text.strip(), zh_text.strip()] if x).strip()
    text_lower = text.lower()
    label_en = resolve_label_en(label_raw)
    flags = {
        "emotion_words_en": _contains_keyword(text_lower, EMOTION_WORDS_EN, lowercase=True) if text else False,
        "emotion_words_cn": _contains_keyword(text, EMOTION_WORDS_CN, lowercase=False) if text else False,
        "emotion_desc_cn": _contains_keyword(text, EMOTION_DESC_CN, lowercase=False) if text else False,
        "label_in_text": bool(
            text
            and (
                (label_en is not None and label_en in text_lower)
                or (label_raw.strip() and label_raw.strip() in text)
            )
        ),
    }
    return flags


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    """读取 XLSX 并规范化成字典列表。

    兼容两类表头：
    1. 显式列名版本，包含 `序号/蒙文/中文/情感类别/...`；
    2. 老式无表头或弱表头版本，只能按列位置猜测字段含义。
    """

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for reading XLSX metadata. Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []

    header = [str(x).strip() if x is not None else "" for x in header_row]
    expected = {
        "序号",
        "情感类别",
        "蒙文",
        "中文",
        "情感强度",
        "强度",
        "intensity",
        "Intensity",
        "音频编号",
        "视频编号",
    }
    has_header = any(h in expected for h in header)

    rows: list[dict[str, Any]] = []
    if has_header:
        col_to_idx = {name: i for i, name in enumerate(header) if name}

        def _get_value(record: tuple[Any, ...], name: str) -> Any:
            """按列名安全读取单元格，越界或缺列时返回 `None`。"""

            idx = col_to_idx.get(name)
            if idx is None or idx >= len(record):
                return None
            value = record[idx]
            if isinstance(value, (str, int, float, bool)) or value is None:
                return value
            return str(value)

        for record in it:
            intensity = None
            for key in ("情感强度", "强度", "intensity", "Intensity"):
                intensity = _get_value(record, key)
                if intensity is not None and str(intensity).strip():
                    break
            rows.append(
                {
                    "序号": _get_value(record, "序号"),
                    "情感类别": _get_value(record, "情感类别"),
                    "蒙文": _get_value(record, "蒙文"),
                    "中文": _get_value(record, "中文"),
                    "情感强度": intensity,
                    "音频编号": _get_value(record, "音频编号"),
                    "视频编号": _get_value(record, "视频编号"),
                }
            )
        return rows

    for record in [header_row, *it]:
        seq = record[0] if len(record) > 0 else None
        mn = record[1] if len(record) > 1 else None
        col2 = record[2] if len(record) > 2 else None
        col3 = record[3] if len(record) > 3 else None
        col4 = record[4] if len(record) > 4 else None

        intensity = None
        zh = None
        label = None
        if col4 is not None and resolve_label_en(col4) is not None:
            intensity = parse_intensity(col2)
            zh = col3
            label = col4
        else:
            zh = col2
            label = col3

        rows.append({"序号": seq, "蒙文": mn, "中文": zh, "情感类别": label, "情感强度": intensity})
    return rows


def resolve_paths_for_seq(data_root: Path, label_en: str, seq: str) -> tuple[Path | None, Path | None]:
    """根据标签目录约定或全局搜索，为样本定位视频和音频路径。"""

    video = data_root / label_en / f"{seq}.mp4"
    audio = data_root / label_en / f"{label_en}_audio" / f"{seq}.wav"
    if video.exists() and audio.exists():
        return video, audio

    video_matches = sorted(data_root.rglob(f"{seq}.mp4"))
    audio_matches = sorted(data_root.rglob(f"{seq}.wav"))
    resolved_video = video_matches[0] if len(video_matches) == 1 else None
    resolved_audio = audio_matches[0] if len(audio_matches) == 1 else None
    return resolved_video, resolved_audio


def manifest_sha256(path: Path) -> str:
    """计算 manifest 文件内容的 SHA256，用于实验可复现追踪。"""

    return hashlib.sha256(path.read_bytes()).hexdigest()


def _summarize_manifest_items(items: list[dict[str, Any]]) -> dict[str, Any]:
    """从 manifest 条目列表生成数据摘要。

    这个摘要会被直接写进 manifest，后续训练/推理都可以快速看到：
    - usable/raw_usable 数量；
    - 每类标签和 speaker 计数；
    - 文本泄漏统计；
    - 是否存在一类情绪只对应一个 speaker 的严重混淆。
    """

    label_counts = Counter()
    split_counts = Counter()
    speaker_counts = Counter()
    text_cue_counts = Counter()
    cue_severity_counts = Counter()
    prompt_group_counts = Counter()
    speaker_to_labels: dict[str, set[str]] = defaultdict(set)
    label_to_speakers: dict[str, set[str]] = defaultdict(set)
    usable_count = 0
    raw_usable_count = 0
    for item in items:
        item = _ensure_text_control_fields(item)
        label_en = item.get("label_en")
        speaker_id = item.get("speaker_id", "UNKNOWN")
        split = item.get("split", "excluded")
        split_counts[str(split)] += 1
        if label_en:
            label_counts[str(label_en)] += 1
            speaker_to_labels[str(speaker_id)].add(str(label_en))
            label_to_speakers[str(label_en)].add(str(speaker_id))
        speaker_counts[str(speaker_id)] += 1
        if item.get("is_usable"):
            usable_count += 1
        if item.get("is_raw_usable"):
            raw_usable_count += 1
        if item.get("text_cue_flag"):
            text_cue_counts["text_cue_flag"] += 1
        for key, value in item.get("text_cue_details", {}).items():
            if value:
                text_cue_counts[str(key)] += 1
        cue_severity_counts[str(item.get("cue_severity", "none"))] += 1
        prompt_group_id = str(item.get("prompt_group_id", "") or "")
        if prompt_group_id:
            prompt_group_counts[prompt_group_id] += 1

    labels_single_speaker = [
        label for label, speakers in label_to_speakers.items() if len({s for s in speakers if s != "UNKNOWN"}) == 1
    ]
    labels_multi_speaker = [
        label for label, speakers in label_to_speakers.items() if len({s for s in speakers if s != "UNKNOWN"}) >= 2
    ]

    return {
        "total_rows": len(items),
        "usable_rows": usable_count,
        "raw_usable_rows": raw_usable_count,
        "label_counts": dict(sorted(label_counts.items())),
        "split_counts": dict(sorted(split_counts.items())),
        "speaker_counts": dict(sorted(speaker_counts.items())),
        "speaker_to_labels": {k: sorted(v) for k, v in sorted(speaker_to_labels.items())},
        "label_to_speakers": {k: sorted(v) for k, v in sorted(label_to_speakers.items())},
        "text_cue_counts": dict(sorted(text_cue_counts.items())),
        "cue_severity_counts": dict(sorted(cue_severity_counts.items())),
        "prompt_groups": int(len(prompt_group_counts)),
        "duplicate_prompt_groups": int(sum(1 for _, count in prompt_group_counts.items() if count > 1)),
        "fatal_confound": bool(labels_single_speaker and not labels_multi_speaker),
        "speaker_group_split_feasible": bool(labels_multi_speaker),
    }


def build_split_manifest(
    *,
    data_root: Path,
    xlsx: Path,
    train_split: float = 0.8,
    seed: int = 42,
    split_strategy: str = "stratified_random_by_label",
) -> dict[str, Any]:
    """构建主工程使用的 split manifest。

    处理流程：
    1. 读取 XLSX；
    2. 规范化序号、标签、文本和强度；
    3. 补齐 speaker、文本线索标记、媒体文件路径和可用性；
    4. 按标签做分层 train/val 划分；
    5. 生成摘要和内容级哈希。

    Returns:
        一个可直接序列化为 JSON 的 manifest 字典。
    """

    rows = read_xlsx_rows(xlsx)
    items: list[dict[str, Any]] = []
    usable_indices_by_label: dict[str, list[int]] = defaultdict(list)

    for row in rows:
        seq = normalize_seq(row.get("序号"))
        label_raw = str(row.get("情感类别", "") or "").strip()
        label_en = resolve_label_en(label_raw)
        mn = str(row.get("蒙文", "") or "").strip()
        zh = str(row.get("中文", "") or "").strip()
        intensity = parse_intensity(row.get("情感强度"))
        speaker_id = infer_speaker_id(label_en)
        cue_details = detect_text_cue_flags(mn, zh, label_raw)
        cue_severity = derive_cue_severity(cue_details)
        masked_mn = mask_emotion_cues(mn, label_raw=label_raw, label_en=label_en)
        masked_zh = mask_emotion_cues(zh, label_raw=label_raw, label_en=label_en)
        prompt_group_text = build_prompt_group_text(mn, zh, label_raw=label_raw, label_en=label_en)
        prompt_group_id = build_prompt_group_id(mn, zh, label_raw=label_raw, label_en=label_en)
        video_path = None
        audio_path = None
        if seq and label_en:
            video_path, audio_path = resolve_paths_for_seq(data_root, label_en, seq)
        item = {
            "seq": seq,
            "label_raw": label_raw,
            "label_en": label_en,
            "label_idx": EMOTIONS.index(label_en) if label_en in EMOTIONS else None,
            "mn": mn,
            "zh": zh,
            "masked_mn": masked_mn,
            "masked_zh": masked_zh,
            "normalized_mn": normalize_text_for_grouping(mn),
            "normalized_zh": normalize_text_for_grouping(zh),
            "prompt_group_text": prompt_group_text,
            "prompt_group_id": prompt_group_id,
            "intensity": intensity,
            "speaker_id": speaker_id,
            "video_path": str(video_path) if video_path is not None else None,
            "audio_path": str(audio_path) if audio_path is not None else None,
            "has_video": bool(video_path is not None and video_path.exists()),
            "has_audio": bool(audio_path is not None and audio_path.exists()),
            "has_text": bool(mn or zh),
            "text_cue_flag": any(cue_details.values()),
            "text_cue_details": cue_details,
            "cue_severity": cue_severity,
            "is_usable": bool(seq and label_en),
            "is_raw_usable": bool(seq and label_en and video_path is not None and audio_path is not None),
            "split": "excluded",
        }
        items.append(item)
        if item["is_usable"] and label_en is not None:
            usable_indices_by_label[label_en].append(len(items) - 1)

    rng = random.Random(int(seed))
    for label_en, indices in usable_indices_by_label.items():
        if split_strategy == "stratified_random_by_label":
            shuffled = list(indices)
            rng.shuffle(shuffled)
        else:
            shuffled = sorted(indices, key=lambda idx: items[idx]["seq"])
        cutoff = int(len(shuffled) * float(train_split))
        # 这里按标签分别切分，确保每个情绪类别在 train/val 都尽量保留比例。
        for index in shuffled[:cutoff]:
            items[index]["split"] = "train"
        for index in shuffled[cutoff:]:
            items[index]["split"] = "val"

    summary = _summarize_manifest_items(items)
    manifest = {
        "manifest_version": 1,
        "data_root": str(data_root),
        "xlsx": str(xlsx),
        "seed": int(seed),
        "train_split": float(train_split),
        "split_strategy": split_strategy,
        "summary": summary,
        "items": items,
    }
    manifest["manifest_sha256"] = hashlib.sha256(
        json.dumps(manifest, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return manifest


def load_split_manifest(path: Path) -> dict[str, Any]:
    """读取并做最基本结构校验。"""

    obj = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(obj, dict) or "items" not in obj:
        raise RuntimeError(f"Invalid split manifest: {path}")
    if not isinstance(obj["items"], list):
        raise RuntimeError(f"Invalid split manifest items: {path}")
    return obj


def select_manifest_items(manifest: dict[str, Any], subset: str = "all") -> list[dict[str, Any]]:
    """从 manifest 中选择某个子集的可用条目。"""

    subset = str(subset or "all").strip().lower()
    items = manifest.get("items", [])
    if subset == "all":
        return [item for item in items if item.get("is_usable")]
    return [item for item in items if item.get("is_usable") and str(item.get("split", "")).lower() == subset]


def _group_items_by_key(items: list[dict[str, Any]], group_key: str) -> list[list[dict[str, Any]]]:
    """Pack manifest items into stable groups by a shared key."""

    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        enriched = _ensure_text_control_fields(item)
        key = str(enriched.get(group_key, "") or enriched.get("seq", ""))
        buckets[key].append(enriched)
    groups = list(buckets.values())
    groups.sort(key=lambda bucket: (str(bucket[0].get(group_key, "")), str(bucket[0].get("seq", ""))))
    return groups


def resolve_grouped_cv_splits(
    items: list[dict[str, Any]],
    *,
    label_names: list[str],
    group_key: str = "prompt_group_id",
    requested_splits: int = 5,
    seed: int = 42,
) -> tuple[list[tuple[list[dict[str, Any]], list[dict[str, Any]]]], int]:
    """Create grouped folds with approximate label balance.

    The algorithm is intentionally deterministic and dependency-light:
    - group samples by `group_key`
    - sort by group size / label diversity
    - greedily assign each group to the fold that best improves label balance
    """

    if requested_splits < 2:
        raise RuntimeError("Grouped CV requires at least 2 splits.")
    groups = _group_items_by_key(items, group_key)
    if len(groups) < 2:
        raise RuntimeError(f"Not enough distinct groups for grouped CV: {len(groups)}")

    total_counts = Counter(str(item.get("label_en", "")) for item in items if str(item.get("label_en", "")) in label_names)
    feasible_splits = min(int(requested_splits), len(groups))
    if feasible_splits < 2:
        raise RuntimeError("Grouped CV became infeasible after group counting.")

    rng = random.Random(int(seed))
    shuffled = list(groups)
    rng.shuffle(shuffled)
    shuffled.sort(
        key=lambda bucket: (
            -len(bucket),
            -len({str(item.get("label_en", "")) for item in bucket}),
            str(bucket[0].get(group_key, "")),
        )
    )

    fold_groups: list[list[list[dict[str, Any]]]] = [[] for _ in range(feasible_splits)]
    fold_counts: list[Counter[str]] = [Counter() for _ in range(feasible_splits)]
    fold_sizes: list[int] = [0 for _ in range(feasible_splits)]
    target_per_fold = {label: total_counts.get(label, 0) / float(feasible_splits) for label in label_names}

    for group in shuffled:
        group_counts = Counter(str(item.get("label_en", "")) for item in group if str(item.get("label_en", "")) in label_names)
        best_fold = 0
        best_score: tuple[float, int, int] | None = None
        for fold_idx in range(feasible_splits):
            score_load = 0.0
            for label in label_names:
                target = max(float(target_per_fold.get(label, 0.0)), 1.0)
                after = fold_counts[fold_idx].get(label, 0) + group_counts.get(label, 0)
                # Lower is better: assign the next group to the fold that is
                # currently *least filled* for the labels carried by this group.
                score_load += (after / target) ** 2
            score = (score_load, fold_sizes[fold_idx] + len(group), fold_idx)
            if best_score is None or score < best_score:
                best_score = score
                best_fold = fold_idx
        fold_groups[best_fold].append(group)
        fold_counts[best_fold].update(group_counts)
        fold_sizes[best_fold] += len(group)

    folds: list[tuple[list[dict[str, Any]], list[dict[str, Any]]]] = []
    for fold_idx in range(feasible_splits):
        val_items = [item for group in fold_groups[fold_idx] for item in group]
        train_items = [
            item
            for other_idx, other_groups in enumerate(fold_groups)
            if other_idx != fold_idx
            for group in other_groups
            for item in group
        ]
        if not train_items or not val_items:
            raise RuntimeError("Grouped CV produced an empty train or val fold.")
        train_groups = {str(item.get(group_key, "")) for item in train_items}
        val_groups = {str(item.get(group_key, "")) for item in val_items}
        overlap = train_groups & val_groups
        if overlap:
            preview = ", ".join(sorted(list(overlap))[:10])
            raise RuntimeError(f"Grouped CV leakage detected: train/val share {group_key}. Example(s): {preview}")
        folds.append((train_items, val_items))
    return folds, feasible_splits


def build_manifest_from_split_items(
    base_manifest: dict[str, Any],
    *,
    train_items: list[dict[str, Any]],
    val_items: list[dict[str, Any]],
    split_strategy: str,
    extra_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build a derived manifest with explicit train/val items."""

    train_ids = {str(item.get("seq", "")) for item in train_items}
    val_ids = {str(item.get("seq", "")) for item in val_items}
    if train_ids & val_ids:
        preview = ", ".join(sorted(list(train_ids & val_ids))[:10])
        raise RuntimeError(f"Derived manifest has overlapping train/val seq values. Example(s): {preview}")

    derived_items: list[dict[str, Any]] = []
    for item in base_manifest.get("items", []):
        copied = _ensure_text_control_fields(item)
        seq = str(copied.get("seq", ""))
        if seq in train_ids:
            copied["split"] = "train"
        elif seq in val_ids:
            copied["split"] = "val"
        elif bool(copied.get("is_usable")):
            copied["split"] = "excluded"
        derived_items.append(copied)

    manifest = {
        "manifest_version": 2,
        "data_root": base_manifest.get("data_root"),
        "xlsx": base_manifest.get("xlsx"),
        "seed": base_manifest.get("seed"),
        "train_split": base_manifest.get("train_split"),
        "split_strategy": split_strategy,
        "summary": _summarize_manifest_items(derived_items),
        "items": derived_items,
    }
    if extra_meta:
        manifest["derived"] = dict(extra_meta)
    manifest["manifest_sha256"] = hashlib.sha256(
        json.dumps(manifest, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return manifest
