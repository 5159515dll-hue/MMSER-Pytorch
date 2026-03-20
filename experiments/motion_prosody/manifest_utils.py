from __future__ import annotations

import hashlib
import json
import math
import random
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any


CN_TO_EN = {
    "愤怒": "angry",
    "厌恶": "disgusted",
    "恐惧": "fear",
    "高兴": "happy",
    "快乐": "happy",
    "开心": "happy",
    "中性": "neutral",
    "悲伤": "sad",
    "惊讶": "surprise",
}

EMOTIONS = ["angry", "disgusted", "fear", "happy", "neutral", "sad", "surprise"]

SPEAKER_LABELS_EN = {
    "A": {"surprise", "fear", "angry"},
    "B": {"neutral", "disgusted"},
    "C": {"happy", "sad"},
}

EMOTION_WORDS_EN = ["happy", "angry", "sad", "fear", "disgust", "disgusted", "surprise", "neutral"]
EMOTION_WORDS_CN = ["高兴", "开心", "快乐", "愤怒", "生气", "悲伤", "伤心", "恐惧", "害怕", "厌恶", "惊讶", "中性"]
EMOTION_DESC_CN = [
    "很开心",
    "非常开心",
    "特别开心",
    "很愤怒",
    "非常愤怒",
    "特别愤怒",
    "很生气",
    "非常生气",
    "特别生气",
    "很悲伤",
    "非常悲伤",
    "很害怕",
    "非常害怕",
    "很恐惧",
    "非常恐惧",
    "很惊讶",
    "非常惊讶",
    "很厌恶",
    "非常厌恶",
]


def normalize_seq(raw: Any) -> str:
    if raw is None or isinstance(raw, bool):
        return ""
    if isinstance(raw, float) and math.isnan(raw):
        return ""
    if isinstance(raw, (int, float)):
        try:
            v = float(raw)
        except Exception:
            return str(raw).strip()
        if math.isnan(v):
            return ""
        if abs(v - round(v)) < 1e-6:
            return str(int(round(v)))
        return str(raw).strip()

    try:
        v = float(str(raw))
        if not math.isnan(v) and abs(v - round(v)) < 1e-6:
            return str(int(round(v)))
    except Exception:
        pass

    s = str(raw).strip()
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        return str(int(s[:-2]))
    if s.isdigit():
        return str(int(s))
    return s


def parse_intensity(raw: Any) -> float | None:
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, float) and math.isnan(raw):
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        v = float(s)
    except Exception:
        return None
    if math.isnan(v) or math.isinf(v):
        return None
    return float(v)


def resolve_label_en(raw: Any) -> str | None:
    if raw is None:
        return None
    label = CN_TO_EN.get(str(raw).strip(), str(raw).strip().lower()).strip().lower()
    return label if label in EMOTIONS else None


def infer_speaker_id(label_en: str | None) -> str:
    if not label_en:
        return "UNKNOWN"
    for speaker_id, labels in SPEAKER_LABELS_EN.items():
        if label_en in labels:
            return speaker_id
    return "UNKNOWN"


def _contains_keyword(text: str, keywords: list[str], *, lowercase: bool) -> bool:
    if lowercase:
        haystack = text.lower()
        return any(kw.lower() in haystack for kw in keywords)
    return any(kw in text for kw in keywords)


def detect_text_cue_flags(mn_text: str, zh_text: str, label_raw: str) -> dict[str, bool]:
    text = " ".join(x for x in [mn_text.strip(), zh_text.strip()] if x).strip()
    text_lower = text.lower()
    label_en = resolve_label_en(label_raw)
    flags = {
        "emotion_words_en": _contains_keyword(text_lower, EMOTION_WORDS_EN, lowercase=True) if text else False,
        "emotion_words_cn": _contains_keyword(text, EMOTION_WORDS_CN, lowercase=False) if text else False,
        "emotion_desc_cn": _contains_keyword(text, EMOTION_DESC_CN, lowercase=False) if text else False,
        "label_in_text": bool(
            text
            and (
                (label_en is not None and label_en in text_lower)
                or (label_raw.strip() and label_raw.strip() in text)
            )
        ),
    }
    return flags


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for reading XLSX metadata. Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []

    header = [str(x).strip() if x is not None else "" for x in header_row]
    expected = {
        "序号",
        "情感类别",
        "蒙文",
        "中文",
        "情感强度",
        "强度",
        "intensity",
        "Intensity",
        "音频编号",
        "视频编号",
    }
    has_header = any(h in expected for h in header)

    rows: list[dict[str, Any]] = []
    if has_header:
        col_to_idx = {name: i for i, name in enumerate(header) if name}

        def _get_value(record: tuple[Any, ...], name: str) -> Any:
            idx = col_to_idx.get(name)
            if idx is None or idx >= len(record):
                return None
            value = record[idx]
            if isinstance(value, (str, int, float, bool)) or value is None:
                return value
            return str(value)

        for record in it:
            intensity = None
            for key in ("情感强度", "强度", "intensity", "Intensity"):
                intensity = _get_value(record, key)
                if intensity is not None and str(intensity).strip():
                    break
            rows.append(
                {
                    "序号": _get_value(record, "序号"),
                    "情感类别": _get_value(record, "情感类别"),
                    "蒙文": _get_value(record, "蒙文"),
                    "中文": _get_value(record, "中文"),
                    "情感强度": intensity,
                    "音频编号": _get_value(record, "音频编号"),
                    "视频编号": _get_value(record, "视频编号"),
                }
            )
        return rows

    for record in [header_row, *it]:
        seq = record[0] if len(record) > 0 else None
        mn = record[1] if len(record) > 1 else None
        col2 = record[2] if len(record) > 2 else None
        col3 = record[3] if len(record) > 3 else None
        col4 = record[4] if len(record) > 4 else None

        intensity = None
        zh = None
        label = None
        if col4 is not None and resolve_label_en(col4) is not None:
            intensity = parse_intensity(col2)
            zh = col3
            label = col4
        else:
            zh = col2
            label = col3

        rows.append({"序号": seq, "蒙文": mn, "中文": zh, "情感类别": label, "情感强度": intensity})
    return rows


def resolve_paths_for_seq(data_root: Path, label_en: str, seq: str) -> tuple[Path | None, Path | None]:
    video = data_root / label_en / f"{seq}.mp4"
    audio = data_root / label_en / f"{label_en}_audio" / f"{seq}.wav"
    if video.exists() and audio.exists():
        return video, audio

    video_matches = sorted(data_root.rglob(f"{seq}.mp4"))
    audio_matches = sorted(data_root.rglob(f"{seq}.wav"))
    resolved_video = video_matches[0] if len(video_matches) == 1 else None
    resolved_audio = audio_matches[0] if len(audio_matches) == 1 else None
    return resolved_video, resolved_audio


def manifest_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _summarize_manifest_items(items: list[dict[str, Any]]) -> dict[str, Any]:
    label_counts = Counter()
    split_counts = Counter()
    speaker_counts = Counter()
    text_cue_counts = Counter()
    speaker_to_labels: dict[str, set[str]] = defaultdict(set)
    label_to_speakers: dict[str, set[str]] = defaultdict(set)
    usable_count = 0
    raw_usable_count = 0
    for item in items:
        label_en = item.get("label_en")
        speaker_id = item.get("speaker_id", "UNKNOWN")
        split = item.get("split", "excluded")
        split_counts[str(split)] += 1
        if label_en:
            label_counts[str(label_en)] += 1
            speaker_to_labels[str(speaker_id)].add(str(label_en))
            label_to_speakers[str(label_en)].add(str(speaker_id))
        speaker_counts[str(speaker_id)] += 1
        if item.get("is_usable"):
            usable_count += 1
        if item.get("is_raw_usable"):
            raw_usable_count += 1
        if item.get("text_cue_flag"):
            text_cue_counts["text_cue_flag"] += 1
        for key, value in item.get("text_cue_details", {}).items():
            if value:
                text_cue_counts[str(key)] += 1

    labels_single_speaker = [
        label for label, speakers in label_to_speakers.items() if len({s for s in speakers if s != "UNKNOWN"}) == 1
    ]
    labels_multi_speaker = [
        label for label, speakers in label_to_speakers.items() if len({s for s in speakers if s != "UNKNOWN"}) >= 2
    ]

    return {
        "total_rows": len(items),
        "usable_rows": usable_count,
        "raw_usable_rows": raw_usable_count,
        "label_counts": dict(sorted(label_counts.items())),
        "split_counts": dict(sorted(split_counts.items())),
        "speaker_counts": dict(sorted(speaker_counts.items())),
        "speaker_to_labels": {k: sorted(v) for k, v in sorted(speaker_to_labels.items())},
        "label_to_speakers": {k: sorted(v) for k, v in sorted(label_to_speakers.items())},
        "text_cue_counts": dict(sorted(text_cue_counts.items())),
        "fatal_confound": bool(labels_single_speaker and not labels_multi_speaker),
        "speaker_group_split_feasible": bool(labels_multi_speaker),
    }


def build_split_manifest(
    *,
    data_root: Path,
    xlsx: Path,
    train_split: float = 0.8,
    seed: int = 42,
    split_strategy: str = "stratified_random_by_label",
) -> dict[str, Any]:
    rows = read_xlsx_rows(xlsx)
    items: list[dict[str, Any]] = []
    usable_indices_by_label: dict[str, list[int]] = defaultdict(list)

    for row in rows:
        seq = normalize_seq(row.get("序号"))
        label_raw = str(row.get("情感类别", "") or "").strip()
        label_en = resolve_label_en(label_raw)
        mn = str(row.get("蒙文", "") or "").strip()
        zh = str(row.get("中文", "") or "").strip()
        intensity = parse_intensity(row.get("情感强度"))
        speaker_id = infer_speaker_id(label_en)
        cue_details = detect_text_cue_flags(mn, zh, label_raw)
        video_path = None
        audio_path = None
        if seq and label_en:
            video_path, audio_path = resolve_paths_for_seq(data_root, label_en, seq)
        item = {
            "seq": seq,
            "label_raw": label_raw,
            "label_en": label_en,
            "label_idx": EMOTIONS.index(label_en) if label_en in EMOTIONS else None,
            "mn": mn,
            "zh": zh,
            "intensity": intensity,
            "speaker_id": speaker_id,
            "video_path": str(video_path) if video_path is not None else None,
            "audio_path": str(audio_path) if audio_path is not None else None,
            "has_video": bool(video_path is not None and video_path.exists()),
            "has_audio": bool(audio_path is not None and audio_path.exists()),
            "has_text": bool(mn or zh),
            "text_cue_flag": any(cue_details.values()),
            "text_cue_details": cue_details,
            "is_usable": bool(seq and label_en),
            "is_raw_usable": bool(seq and label_en and video_path is not None and audio_path is not None),
            "split": "excluded",
        }
        items.append(item)
        if item["is_usable"] and label_en is not None:
            usable_indices_by_label[label_en].append(len(items) - 1)

    rng = random.Random(int(seed))
    for label_en, indices in usable_indices_by_label.items():
        if split_strategy == "stratified_random_by_label":
            shuffled = list(indices)
            rng.shuffle(shuffled)
        else:
            shuffled = sorted(indices, key=lambda idx: items[idx]["seq"])
        cutoff = int(len(shuffled) * float(train_split))
        for index in shuffled[:cutoff]:
            items[index]["split"] = "train"
        for index in shuffled[cutoff:]:
            items[index]["split"] = "val"

    summary = _summarize_manifest_items(items)
    manifest = {
        "manifest_version": 1,
        "data_root": str(data_root),
        "xlsx": str(xlsx),
        "seed": int(seed),
        "train_split": float(train_split),
        "split_strategy": split_strategy,
        "summary": summary,
        "items": items,
    }
    manifest["manifest_sha256"] = hashlib.sha256(
        json.dumps(manifest, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return manifest


def load_split_manifest(path: Path) -> dict[str, Any]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(obj, dict) or "items" not in obj:
        raise RuntimeError(f"Invalid split manifest: {path}")
    if not isinstance(obj["items"], list):
        raise RuntimeError(f"Invalid split manifest items: {path}")
    return obj


def select_manifest_items(manifest: dict[str, Any], subset: str = "all") -> list[dict[str, Any]]:
    subset = str(subset or "all").strip().lower()
    items = manifest.get("items", [])
    if subset == "all":
        return [item for item in items if item.get("is_usable")]
    return [item for item in items if item.get("is_usable") and str(item.get("split", "")).lower() == subset]
